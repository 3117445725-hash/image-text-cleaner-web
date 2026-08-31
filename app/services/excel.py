from __future__ import annotations

import csv
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.config import MAX_URLS_PER_JOB
from app.services.downloader import ImageDownloader
from app.services.matcher import match_text
from app.services.ocr import ocr_service

URL_RE = re.compile(r"https?://[^\s<>\"',，;；|｜]+", re.IGNORECASE)
HYPERLINK_FORMULA_RE = re.compile(
    r'^=HYPERLINK\(\s*"(?P<url>https?://[^"]+)"\s*[;,]\s*"(?P<label>[^"]*)"\s*\)$',
    re.IGNORECASE,
)
MAX_ARCHIVE_FILES = 20_000
MAX_UNCOMPRESSED_BYTES = 500 * 1024 * 1024


@dataclass
class LinkOccurrence:
    sheet: str
    cell: str
    url: str
    kind: str


@dataclass
class LinkResult:
    url: str
    status: str
    matched: bool
    matched_keywords: tuple[str, ...]
    ocr_text: str
    average_score: float | None
    error: str


def _validate_workbook_archive(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_FILES:
                raise ValueError("Excel 文件内部文件数量异常，已拒绝处理")
            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("Excel 解压后体积过大，已拒绝处理")
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            names = {info.filename for info in infos}
            if not required.issubset(names):
                raise ValueError("文件不是有效的 Excel 工作簿")
    except zipfile.BadZipFile as exc:
        raise ValueError("文件不是有效的 Excel 工作簿") from exc


def _extract_occurrences(workbook) -> list[LinkOccurrence]:
    found: list[LinkOccurrence] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.hyperlink and cell.hyperlink.target and str(cell.hyperlink.target).lower().startswith(("http://", "https://")):
                    found.append(LinkOccurrence(sheet.title, cell.coordinate, str(cell.hyperlink.target), "hyperlink"))

                value = cell.value
                if isinstance(value, str):
                    formula_match = HYPERLINK_FORMULA_RE.match(value.strip())
                    if formula_match:
                        found.append(LinkOccurrence(sheet.title, cell.coordinate, formula_match.group("url"), "formula"))
                        continue
                    for match in URL_RE.finditer(value):
                        found.append(LinkOccurrence(sheet.title, cell.coordinate, match.group(0).rstrip(".,;，；)）]】"), "value"))
    return found


def _remove_url_from_text(text: str, url: str) -> str | None:
    result = text.replace(url, "")
    result = re.sub(r"[ \t]*([,，;；|｜]+[ \t]*){2,}", r"\1", result)
    result = re.sub(r"^[\s,，;；|｜]+|[\s,，;；|｜]+$", "", result)
    result = re.sub(r"[ \t]{2,}", " ", result).strip()
    return result if result else None


def _csv_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def process_workbook(
    input_path: Path,
    output_path: Path,
    log_path: Path,
    *,
    keywords: list[str],
    match_mode: str,
    case_sensitive: bool,
    enhanced_ocr: bool,
    progress_callback: Callable[[int, int, int, int, str], None],
) -> tuple[int, int, int, int]:
    _validate_workbook_archive(input_path)
    keep_vba = input_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(input_path, data_only=False, keep_links=True, keep_vba=keep_vba)
    occurrences = _extract_occurrences(workbook)
    unique_urls = list(dict.fromkeys(item.url for item in occurrences))

    if len(unique_urls) > MAX_URLS_PER_JOB:
        raise ValueError(f"图片链接数量为 {len(unique_urls)}，超过单次 {MAX_URLS_PER_JOB} 条限制")

    results: dict[str, LinkResult] = {}
    downloader = ImageDownloader()
    matched_count = 0
    failed_count = 0
    try:
        total = len(unique_urls)
        for index, url in enumerate(unique_urls, start=1):
            try:
                downloaded = downloader.fetch(url)
                ocr = ocr_service.recognize(downloaded.content, enhanced=enhanced_ocr)
                match = match_text(ocr.text, keywords, mode=match_mode, case_sensitive=case_sensitive)
                if match.matched:
                    matched_count += 1
                results[url] = LinkResult(
                    url=url,
                    status="matched" if match.matched else "not_matched",
                    matched=match.matched,
                    matched_keywords=match.matched_keywords,
                    ocr_text=ocr.text,
                    average_score=ocr.average_score,
                    error="",
                )
            except Exception as exc:  # noqa: BLE001 - every failed URL must be logged and retained
                failed_count += 1
                results[url] = LinkResult(
                    url=url,
                    status="failed",
                    matched=False,
                    matched_keywords=(),
                    ocr_text="",
                    average_score=None,
                    error=str(exc),
                )
            progress_callback(index, total, matched_count, failed_count, url)
    finally:
        downloader.close()

    changed_cells: set[tuple[str, str]] = set()
    for occurrence in occurrences:
        result = results[occurrence.url]
        if not result.matched:
            continue
        sheet = workbook[occurrence.sheet]
        cell = sheet[occurrence.cell]
        if occurrence.kind == "hyperlink":
            cell.hyperlink = None
            if isinstance(cell.value, str) and cell.value.strip() == occurrence.url:
                cell.value = None
        elif occurrence.kind == "formula":
            cell.value = None
        else:
            if isinstance(cell.value, str):
                cell.value = _remove_url_from_text(cell.value, occurrence.url)
        changed_cells.add((occurrence.sheet, occurrence.cell))

    workbook.save(output_path)

    with log_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "工作表", "单元格", "图片链接", "链接类型", "处理状态", "是否命中",
            "命中关键词", "OCR识别全文", "平均置信度", "错误信息",
        ])
        for occurrence in occurrences:
            result = results[occurrence.url]
            writer.writerow([
                _csv_safe(occurrence.sheet),
                occurrence.cell,
                occurrence.url,
                occurrence.kind,
                result.status,
                "是" if result.matched else "否",
                _csv_safe(" | ".join(result.matched_keywords)),
                _csv_safe(result.ocr_text),
                "" if result.average_score is None else f"{result.average_score:.4f}",
                _csv_safe(result.error),
            ])

    return len(unique_urls), matched_count, failed_count, len(changed_cells)
